import random
import math
import pylab as pl
import numpy as np
from matplotlib.colors import ListedColormap

#Чтение данных из exel
import openpyxl
wookbook = openpyxl.load_workbook("dataset.xlsx")
worksheet = wookbook.active
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")
    print('')

#Train data generator
def generateData (numberOfClassEl, numberOfClasses):
    data = []
    for classNum in range(numberOfClasses):
        #Choose random center of 2-dimensional gaussian
        centerX, centerY = random.random()*5.0, random.random()*5.0
        #Choose numberOfClassEl random nodes with RMS=0.5
        for rowNum in range(numberOfClassEl):
            data.append([ [random.gauss(centerX,0.5), random.gauss(centerY,0.5)], classNum])
    return data

def showData (nClasses, nItemsInClass):
    trainData      = generateData (nItemsInClass, nClasses)
    classColormap  = ListedColormap(['#FF0000', '#00FF00', '#FFFFFF'])
    pl.scatter([trainData[i][0][0] for i in range(len(trainData))],
               [trainData[i][0][1] for i in range(len(trainData))],
               c=[trainData[i][1] for i in range(len(trainData))],
               cmap=classColormap)
    pl.show()   
showData (3, 40)

#Separate N data elements in two parts:
#	test data with N*testPercent elements
#	train_data with N*(1.0 - testPercent) elements
def splitTrainTest (data, testPercent):
    trainData = []
    testData  = []
    for row in data:
        if random.random() < testPercent:
            testData.append(row)
        else:
            trainData.append(row)
    return trainData, testData	
#Main classification procedure
def classifyKNN (trainData, testData, k, numberOfClasses):
    #Euclidean distance between 2-dimensional point
    def dist (a, b):
        return math.sqrt((a[0] - b[0])**2 + (a[1] - b[1])**2)
    testLabels = []	
    for testPoint in testData:
        #Claculate distances between test point and all of the train points
        testDist = [ [dist(testPoint, trainData[i][0]), trainData[i][1]] for i in range(len(trainData))]
        #How many points of each class among nearest K
        stat = [0 for i in range(numberOfClasses)]
        for d in sorted(testDist)[0:k]:
            stat[d[1]] += 1
        #Assign a class with the most number of occurences among K nearest neighbours
        testLabels.append( sorted(zip(stat, range(numberOfClasses)), reverse=True)[0][1] )
    return testLabels
